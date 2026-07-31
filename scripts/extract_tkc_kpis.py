"""
TKC形式・科目別月次推移表から月次KPI（売上高／変動費計／限界利益率／人件費／労働分配率）を
抽出するヘルパー。計算ルールは ../references/calc-rules.md を参照。

このスクリプトは「ライブラリ」として使う想定。会社によって科目コードの体系や、ヘッダー行の
位置が微妙に違うことがあるので、まず対話的に構造を確認してから
（例: read_workbook_preview で先頭数行・科目一覧を見る）、
月ヘッダーの行番号やデータ開始行を渡して呼び出すこと。決め打ちで実行しない。

使い方の例（Pythonから import して使う）:

    import openpyxl
    from extract_tkc_kpis import find_month_columns, compute_monthly_kpis

    wb = openpyxl.load_workbook("科目残高.xlsx", data_only=True)
    ws = wb.active
    months = find_month_columns(ws, header_row=2)   # {"2026-05": 52, ...} (借方列のインデックス)
    kpis = compute_monthly_kpis(ws, months)          # {"2026-05": {"sales": ..., ...}, ...}

CLIとしても使える（ざっくり確認したいとき）:

    python extract_tkc_kpis.py 科目残高.xlsx
"""
import re
import sys
from datetime import datetime, timedelta

DEFAULT_REVENUE_CODE_RANGE = (4111, 4115)
DEFAULT_LABOR_CODES = {
    5431, 5432, 5433, 5434, 5435, 5438,   # 製造原価側：賃金・賞与・雑給・法定福利費・厚生費・退職金
    6111, 6211, 6212, 6213, 6312, 6226,   # 販管費側：給与手当・役員報酬・事務員給与・従業員賞与・法定福利費・厚生費
}
VARIABLE_FLAG = "Ｖ"  # 全角V。元データの表記に合わせて呼び出し側で上書き可能


def read_workbook_preview(path, sheet=None, rows=6, cols=12):
    """最初に構造を目視確認するための軽量プレビュー。会社ごとにヘッダー位置が違う可能性が
    あるため、決め打ちで extract に進まず、まずこれで科目コード列・固変フラグ列・月ヘッダーの
    位置を確認すること。"""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    preview = []
    for r in range(1, rows + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, cols + 1)]
        preview.append(row)
    return ws.title, preview


_ERA_RE = re.compile(r"[RrＲ令和]\s*(\d+)\D+(\d{1,2})")


def _era_to_ad_year(era_year, era="令和"):
    # 令和1年 = 2019年
    return era_year + 2018


def find_month_columns(ws, header_row=2, code_col=1, max_col=None):
    """月ヘッダー行（例: 'R 7. 1' や '令和 7年 1月'）を左から走査し、
    {"YYYY-MM": debit_col_index} を返す。3列1組（借方・貸方・残高）を仮定しているので、
    貸方はdebit_col+1、残高はdebit_col+2。

    ヘッダーの書式が上記と違う場合（和暦でなく西暦、月名表記など）は、この関数を使わず
    月ごとの列番号を手作業で確認してdictを直接組み立てる方が安全。
    """
    max_col = max_col or ws.max_column
    months = {}
    c = code_col + 1
    while c <= max_col:
        v = ws.cell(row=header_row, column=c).value
        if v:
            m = _ERA_RE.search(str(v))
            if m:
                year = _era_to_ad_year(int(m.group(1)))
                month = int(m.group(2))
                months[f"{year:04d}-{month:02d}"] = c
                c += 3
                continue
        c += 1
    return months


def compute_monthly_kpis(
    ws,
    month_debit_cols,
    revenue_code_range=DEFAULT_REVENUE_CODE_RANGE,
    labor_codes=DEFAULT_LABOR_CODES,
    variable_flag=VARIABLE_FLAG,
    data_start_row=4,
    data_end_row=None,
    code_col=1,
    flag_col=2,
):
    """month_debit_cols: {"YYYY-MM": debit_col_index} (find_month_columnsの出力)
    戻り値: {"YYYY-MM": {"sales":円, "variable_cost":円, "labor":円, "margin":円,
                          "margin_pct":0-100, "labor_ratio":0-100}}
    金額はすべて円建て（呼び出し側で /1000 して千円に変換すること）。
    """
    data_end_row = data_end_row or ws.max_row
    rows = []
    for r in range(data_start_row, data_end_row + 1):
        code = ws.cell(row=r, column=code_col).value
        flag = ws.cell(row=r, column=flag_col).value
        if code is None:
            continue
        rows.append((r, code, flag))

    lo, hi = revenue_code_range
    result = {}
    for label, dcol in month_debit_cols.items():
        ccol = dcol + 1
        sales = variable_cost = labor = 0
        for (r, code, flag) in rows:
            debit = ws.cell(row=r, column=dcol).value or 0
            credit = ws.cell(row=r, column=ccol).value or 0
            if isinstance(code, (int, float)) and lo <= code <= hi:
                sales += credit - debit
            if flag and variable_flag in str(flag):
                variable_cost += debit - credit
            if code in labor_codes:
                labor += debit - credit
        margin = sales - variable_cost
        result[label] = {
            "sales": sales,
            "variable_cost": variable_cost,
            "labor": labor,
            "margin": margin,
            "margin_pct": (margin / sales * 100) if sales else None,
            "labor_ratio": (labor / margin * 100) if margin else None,
        }
    return result


def parse_order_history(ws, date_row=2, order_row=3, sales_row=4, start_col=3, end_col=None):
    """「受注履歴」的なシート（1行=日付のシリアル値、次の行が受注高、その次が売上高）を
    {"YYYY-MM": (受注高, 売上高)} に変換する。列の並びは会社ごとに異なるので、
    date_row/order_row/sales_row/start_col は必ず実データを見てから指定すること。"""
    end_col = end_col or ws.max_column
    out = {}
    for c in range(start_col, end_col + 1):
        d = ws.cell(row=date_row, column=c).value
        j = ws.cell(row=order_row, column=c).value
        u = ws.cell(row=sales_row, column=c).value
        if d is None:
            continue
        if isinstance(d, (int, float)):
            dt = datetime(1899, 12, 30) + timedelta(days=d)
            label = dt.strftime("%Y-%m")
        else:
            label = str(d)
        out[label] = (j, u)
    return out


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("usage: python extract_tkc_kpis.py <科目残高.xlsx> [sheet_name]")
        sys.exit(1)
    path = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv) > 2 else None
    title, preview = read_workbook_preview(path, sheet)
    print(f"sheet: {title}")
    for row in preview:
        print(row)
